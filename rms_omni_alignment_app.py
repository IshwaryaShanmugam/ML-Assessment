"""
RMS vs OMNI Alignment Validator
Windows desktop application.

Requirements:
    pip install pandas openpyxl

Run:
    python rms_omni_alignment_app.py
"""

from __future__ import annotations

import os
import re
import threading
import traceback
from collections import Counter
from pathlib import Path
from typing import Dict, List, Tuple, Optional, Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import tkinter as tk
from tkinter import filedialog, messagebox
from tkinter import ttk


# -----------------------------
# Configuration
# -----------------------------

INPUT_OMNI_SHEET = "OMNI IV"
INPUT_MAC_SHEET = "MAC"

OUTPUT_SHEETS = ["OMNI IV", "MAC", "Analysis", "Rule"]
OUTPUT_FILENAME = "RMS_vs_OMNI_Alignment_Result.xlsx"

ANALYSIS_COLUMNS = [
    "OMNI ID",
    "RMS ID",
    "Item Scope",
    "OMNI Value",
    "RMS Value",
    "Alignment Statement",
    "Validation Status",
    "Unique RMS Value",
    "Derived Brand High",
    "#US LOC BRAND",
    "Coding / Combination",
    "Rule Level",
    "Impacted Nankey Count",
    "Unique RMS Count",
    "Final Status",
    "Remarks",
]

RULE_COLUMNS = [
    "OMNI ID",
    "RMS ID",
    "Rule Level",
    "Derived Brand High",
    "#US LOC BRAND",
    "Coding / Combination",
    "RMS Value",
    "Impacted Nankey Count",
    "Unique RMS Count",
    "Status",
    "Recommended OMNI Value",
    "Remarks",
]

# These are business exclusions from the supplied process.
EXCLUDED_RMS_PATTERNS = [
    re.compile(r"\bPRIVATE\s+LABEL\b", re.I),
    re.compile(r"\bAO\s+BRANDS?\b", re.I),
]

DERIVED_BRAND_CANDIDATES = [
    "#US LOC DERIVED BRAND HIGH [84561]",
    "#US LOC DERIVED BRAND HIGH",
]

BRAND_CANDIDATES = [
    "#US LOC BRAND [71177]",
    "#US LOC BRAND",
]

NANKEY_CANDIDATES = ["nan_key", "nankey", "nan key"]

CODING_HINTS = (
    "coding",
    "code",
    "combination",
    "combo",
)


# -----------------------------
# Generic helpers
# -----------------------------

def norm_header(value: Any) -> str:
    """Normalize a header for tolerant matching."""
    if value is None:
        return ""
    text = str(value).strip()
    # Remove a leading Excel formula marker from formula-generated headers.
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text)
    return text.casefold()


def clean_value(value: Any) -> str:
    """Return a stable, human-readable value while preserving original text."""
    if value is None:
        return ""
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value).strip())


def compare_value(value: Any) -> str:
    """Canonical value used for equality/grouping."""
    return clean_value(value).casefold()


def is_excluded_rms(value: Any) -> bool:
    text = clean_value(value)
    if not text:
        return False
    return any(pattern.search(text) for pattern in EXCLUDED_RMS_PATTERNS)


def find_unique_column(columns: List[Any], candidates: List[str], label: str) -> str:
    """Find one column by normalized candidate names."""
    lookup = {}
    duplicates = set()
    for col in columns:
        key = norm_header(col)
        if not key:
            continue
        if key in lookup:
            duplicates.add(key)
        else:
            lookup[key] = col

    for candidate in candidates:
        key = norm_header(candidate)
        if key in duplicates:
            raise ValueError(
                f"Duplicate/ambiguous {label} column detected: '{candidate}'. "
                "Please make the header unique."
            )
        if key in lookup:
            return lookup[key]

    raise ValueError(
        f"Required {label} column is missing. "
        f"Expected one of: {', '.join(candidates)}"
    )


def detect_column(columns: List[Any], candidates: List[str], label: str) -> Optional[str]:
    try:
        return find_unique_column(columns, candidates, label)
    except ValueError:
        return None


def detect_coding_columns(columns: List[Any]) -> List[str]:
    """
    Detect actual coding/combination fields from the workbook.
    The application deliberately does not guess arbitrary business fields.
    """
    result = []
    seen = set()

    for col in columns:
        h = norm_header(col)
        if not h:
            continue
        if ("coding" in h or "combination" in h or "combo" in h or re.search(r"\bcode\b", h)):

            if h not in seen:
                result.append(col)
                seen.add(h)

    return result


def make_unique_headers(raw_headers: List[Any]) -> List[str]:
    """
    Create DataFrame-safe headers while retaining original header text.
    Blank headers are given internal names.
    """
    counts = Counter()
    output = []

    for idx, header in enumerate(raw_headers, start=1):
        base = clean_value(header)
        if not base:
            base = f"__BLANK_COLUMN_{idx}__"

        counts[base] += 1
        if counts[base] == 1:
            output.append(base)
        else:
            output.append(f"{base}__DUPLICATE_{counts[base]}")

    return output


def read_omni_sheet(path: str) -> Tuple[pd.DataFrame, List[str], List[str]]:
    """
    Read OMNI IV using row 2 as the header row, matching the supplied workbook.
    Returns:
        dataframe, original headers, generated unique dataframe headers
    """
    # Read raw first so headers can be preserved exactly.
    raw = pd.read_excel(path, sheet_name=INPUT_OMNI_SHEET, header=None, dtype=object)
    if raw.shape[0] < 2:
        raise ValueError("OMNI IV does not contain the expected header row.")

    original_headers = [clean_value(x) for x in raw.iloc[1].tolist()]
    unique_headers = make_unique_headers(original_headers)

    data = raw.iloc[2:].copy()
    data.columns = unique_headers
    data = data.dropna(how="all").reset_index(drop=True)

    return data, original_headers, unique_headers


def read_mac_sheet(path: str) -> pd.DataFrame:
    mac = pd.read_excel(path, sheet_name=INPUT_MAC_SHEET, dtype=object)
    if mac.empty:
        raise ValueError("MAC sheet is empty.")

    mac.columns = make_unique_headers(list(mac.columns))

    # Locate the required mapping headers.
    omni_col = find_unique_column(list(mac.columns), ["OMNI ID"], "MAC OMNI ID")
    rms_col = find_unique_column(list(mac.columns), ["RMS ID"], "MAC RMS ID")

    mac = mac[[omni_col, rms_col]].copy()
    mac.columns = ["OMNI ID", "RMS ID"]
    mac = mac.dropna(how="all")

    mappings = []
    seen = set()
    for _, row in mac.iterrows():
        omni = clean_value(row["OMNI ID"])
        rms = clean_value(row["RMS ID"])
        if not omni and not rms:
            continue
        if not omni or not rms:
            raise ValueError(
                "MAC contains a row with a missing OMNI ID or RMS ID."
            )

        key = (norm_header(omni), norm_header(rms))
        if key in seen:
            continue
        seen.add(key)
        mappings.append((omni, rms))

    if not mappings:
        raise ValueError("No valid OMNI ID → RMS ID mappings were found in MAC.")

    return pd.DataFrame(mappings, columns=["OMNI ID", "RMS ID"])


def validate_headers(df: pd.DataFrame, omni_id: str, rms_id: str,
                     derived_col: str, brand_col: str, nankey_col: str) -> None:
    required = {
        "OMNI ID": omni_id,
        "RMS ID": rms_id,
        "Derived Brand High": derived_col,
        "#US LOC BRAND": brand_col,
        "Nankey": nankey_col,
    }

    missing = [label for label, col in required.items() if col not in df.columns]
    if missing:
        raise ValueError(
            "Required analysis columns are missing from OMNI IV: "
            + ", ".join(missing)
        )


# -----------------------------
# Core validation engine
# -----------------------------

def value_set(series: pd.Series, exclude_business_values: bool = False) -> List[str]:
    vals = []
    seen = set()
    for x in series:
        text = clean_value(x)
        if not text:
            continue
        if exclude_business_values and is_excluded_rms(text):
            continue
        key = compare_value(text)
        if key not in seen:
            seen.add(key)
            vals.append(text)
    return vals


def group_info(group: pd.DataFrame, rms_col: str, nankey_col: str) -> Tuple[List[str], int]:
    rms_values = value_set(group[rms_col], exclude_business_values=True)
    nankey_count = group[nankey_col].map(clean_value).replace("", pd.NA).dropna().nunique()
    return rms_values, int(nankey_count)


def coding_label(group: pd.DataFrame, coding_cols: List[str]) -> str:
    if not coding_cols:
        return ""

    pieces = []
    for col in coding_cols:
        vals = value_set(group[col], exclude_business_values=False)
        if vals:
            pieces.append(f"{col}: {' | '.join(vals)}")
        else:
            pieces.append(f"{col}: <blank>")
    return " ; ".join(pieces)


def coding_key(group: pd.DataFrame, coding_cols: List[str]) -> pd.Series:
    """
    Return a canonical combination key for detected coding/combination fields.
    """
    if not coding_cols:
        return pd.Series([""] * len(group), index=group.index)

    result = []
    for idx in group.index:
        parts = [compare_value(group.loc[idx, col]) for col in coding_cols]
        result.append("||".join(parts))
    return pd.Series(result, index=group.index)


def build_analysis_and_rules(df: pd.DataFrame, mappings: pd.DataFrame):
    item_scope_col = find_unique_column(
        list(df.columns), ["Item_Scope", "Item Scope"], "Item Scope"
    )
    derived_col = detect_column(df.columns.tolist(), DERIVED_BRAND_CANDIDATES,
                                 "Derived Brand High")
    brand_col = detect_column(df.columns.tolist(), BRAND_CANDIDATES,
                              "#US LOC BRAND")
    nankey_col = detect_column(df.columns.tolist(), NANKEY_CANDIDATES, "Nankey")

    if not derived_col:
        raise ValueError("The Derived Brand High column is missing from OMNI IV.")
    if not brand_col:
        raise ValueError("The #US LOC BRAND column is missing from OMNI IV.")
    if not nankey_col:
        raise ValueError("The Nankey column (for example nan_key) is missing.")

    coding_cols = detect_coding_columns(df.columns.tolist())

    # Scope filter. Blank scope is not treated as SAME IN BOTH.
    scoped = df[
        df[item_scope_col].map(compare_value).eq("same in both")
    ].copy()

    analysis_rows: List[Dict[str, Any]] = []
    rule_rows: List[Dict[str, Any]] = []

    for _, mapping in mappings.iterrows():
        omni_col = clean_value(mapping["OMNI ID"])
        rms_col = clean_value(mapping["RMS ID"])

        # Mapping must point to actual OMNI IV columns.
        if omni_col not in df.columns or rms_col not in df.columns:
            raise ValueError(
                f"Mapped column does not exist in OMNI IV: "
                f"OMNI='{omni_col}' / RMS='{rms_col}'."
            )

        # Dynamic comparison.
        omni_cmp = scoped[omni_col].map(compare_value)
        rms_cmp = scoped[rms_col].map(compare_value)
        false_mask = omni_cmp.ne(rms_cmp)

        # Treat both blank as aligned; only actual differences are false.
        false_df = scoped.loc[false_mask].copy()

        if false_df.empty:
            continue

        # Exclude business values only for the unique-RMS rule-analysis stage.
        false_df["_RMS_CLEAN"] = false_df[rms_col].map(clean_value)
        eligible = false_df[
            false_df["_RMS_CLEAN"].ne("")
            & ~false_df["_RMS_CLEAN"].map(is_excluded_rms)
        ].copy()

        # If every false RMS value is blank/excluded, preserve the trace as REVIEW.
        if eligible.empty:
            analysis_rows.append({
                "OMNI ID": omni_col,
                "RMS ID": rms_col,
                "Item Scope": "SAME IN BOTH",
                "OMNI Value": "",
                "RMS Value": "",
                "Alignment Statement": f"{omni_col} != {rms_col}",
                "Validation Status": "FALSE",
                "Unique RMS Value": "",
                "Derived Brand High": "",
                "#US LOC BRAND": "",
                "Coding / Combination": "",
                "Rule Level": "REVIEW",
                "Impacted Nankey Count": int(false_df[nankey_col].map(clean_value).replace("", pd.NA).dropna().nunique()),
                "Unique RMS Count": 0,
                "Final Status": "REVIEW",
                "Remarks": "No usable RMS value after excluding PRIVATE LABEL / AO BRAND values.",
            })
            rule_rows.append({
                "OMNI ID": omni_col,
                "RMS ID": rms_col,
                "Rule Level": "REVIEW",
                "Derived Brand High": "",
                "#US LOC BRAND": "",
                "Coding / Combination": "",
                "RMS Value": "",
                "Impacted Nankey Count": int(false_df[nankey_col].map(clean_value).replace("", pd.NA).dropna().nunique()),
                "Unique RMS Count": 0,
                "Status": "REVIEW",
                "Recommended OMNI Value": "",
                "Remarks": "No usable RMS value after excluding PRIVATE LABEL / AO BRAND values.",
            })
            continue

        # Level 1: Derived Brand High -> RMS.
        for derived_value, dgroup in eligible.groupby(
            derived_col, dropna=False, sort=False
        ):
            derived_display = clean_value(derived_value)
            rms_values, nankeys = group_info(dgroup, rms_col, nankey_col)

            if len(rms_values) == 1:
                status = "PASS"
                level = "LEVEL 1"
                remarks = "Derived Brand High uniquely identifies one RMS value."
                brand_display = ""
                coding_display = ""
                recommended = rms_values[0]

                analysis_rows.append({
                    "OMNI ID": omni_col,
                    "RMS ID": rms_col,
                    "Item Scope": "SAME IN BOTH",
                    "OMNI Value": "",
                    "RMS Value": rms_values[0],
                    "Alignment Statement": f"{omni_col} != {rms_col}",
                    "Validation Status": "FALSE",
                    "Unique RMS Value": " | ".join(rms_values),
                    "Derived Brand High": derived_display,
                    "#US LOC BRAND": "",
                    "Coding / Combination": "",
                    "Rule Level": level,
                    "Impacted Nankey Count": nankeys,
                    "Unique RMS Count": len(rms_values),
                    "Final Status": status,
                    "Remarks": remarks,
                })

                rule_rows.append({
                    "OMNI ID": omni_col,
                    "RMS ID": rms_col,
                    "Rule Level": level,
                    "Derived Brand High": derived_display,
                    "#US LOC BRAND": "",
                    "Coding / Combination": "",
                    "RMS Value": rms_values[0],
                    "Impacted Nankey Count": nankeys,
                    "Unique RMS Count": len(rms_values),
                    "Status": status,
                    "Recommended OMNI Value": recommended,
                    "Remarks": remarks,
                })
                continue

            # Level 1 multiple breakout: go to Level 2.
            dgroup = dgroup.copy()
            dgroup["_RMS_CANON"] = dgroup[rms_col].map(compare_value)

            level2_found = False
            for (derived_key, brand_value), bgroup in dgroup.groupby(
                [derived_col, brand_col], dropna=False, sort=False
            ):
                b_rms_values, b_nankeys = group_info(bgroup, rms_col, nankey_col)
                brand_display = clean_value(brand_value)

                if len(b_rms_values) == 1:
                    level2_found = True
                    status = "PASS"
                    level = "LEVEL 2"
                    remarks = (
                        "Derived Brand High has multiple RMS breakouts; "
                        "#US LOC BRAND uniquely resolves the RMS value."
                    )

                    analysis_rows.append({
                        "OMNI ID": omni_col,
                        "RMS ID": rms_col,
                        "Item Scope": "SAME IN BOTH",
                        "OMNI Value": "",
                        "RMS Value": b_rms_values[0],
                        "Alignment Statement": f"{omni_col} != {rms_col}",
                        "Validation Status": "FALSE",
                        "Unique RMS Value": " | ".join(b_rms_values),
                        "Derived Brand High": derived_display,
                        "#US LOC BRAND": brand_display,
                        "Coding / Combination": "",
                        "Rule Level": level,
                        "Impacted Nankey Count": b_nankeys,
                        "Unique RMS Count": len(b_rms_values),
                        "Final Status": status,
                        "Remarks": remarks,
                    })

                    rule_rows.append({
                        "OMNI ID": omni_col,
                        "RMS ID": rms_col,
                        "Rule Level": level,
                        "Derived Brand High": derived_display,
                        "#US LOC BRAND": brand_display,
                        "Coding / Combination": "",
                        "RMS Value": b_rms_values[0],
                        "Impacted Nankey Count": b_nankeys,
                        "Unique RMS Count": len(b_rms_values),
                        "Status": status,
                        "Recommended OMNI Value": b_rms_values[0],
                        "Remarks": remarks,
                    })
                else:
                    # Level 3: use only actual coding/combination columns detected.
                    if not coding_cols:
                        coding_display = ""
                        remarks = (
                            "Level 2 remains ambiguous. No actual coding/combination "
                            "column was detected in the workbook, so no rule was "
                            "generated automatically."
                        )

                        analysis_rows.append({
                            "OMNI ID": omni_col,
                            "RMS ID": rms_col,
                            "Item Scope": "SAME IN BOTH",
                            "OMNI Value": "",
                            "RMS Value": " | ".join(b_rms_values),
                            "Alignment Statement": f"{omni_col} != {rms_col}",
                            "Validation Status": "FALSE",
                            "Unique RMS Value": " | ".join(b_rms_values),
                            "Derived Brand High": derived_display,
                            "#US LOC BRAND": brand_display,
                            "Coding / Combination": coding_display,
                            "Rule Level": "LEVEL 3",
                            "Impacted Nankey Count": b_nankeys,
                            "Unique RMS Count": len(b_rms_values),
                            "Final Status": "REVIEW",
                            "Remarks": remarks,
                        })

                        rule_rows.append({
                            "OMNI ID": omni_col,
                            "RMS ID": rms_col,
                            "Rule Level": "LEVEL 3",
                            "Derived Brand High": derived_display,
                            "#US LOC BRAND": brand_display,
                            "Coding / Combination": "",
                            "RMS Value": " | ".join(b_rms_values),
                            "Impacted Nankey Count": b_nankeys,
                            "Unique RMS Count": len(b_rms_values),
                            "Status": "REVIEW",
                            "Recommended OMNI Value": "",
                            "Remarks": remarks,
                        })
                        continue

                    # Group by the actual detected coding/combination fields.
                    bgroup = bgroup.copy()
                    bgroup["_CODING_KEY"] = coding_key(bgroup, coding_cols)

                    for code_key, cgroup in bgroup.groupby(
                        "_CODING_KEY", dropna=False, sort=False
                    ):
                        c_rms_values, c_nankeys = group_info(
                            cgroup, rms_col, nankey_col
                        )
                        c_display = coding_label(cgroup, coding_cols)

                        if len(c_rms_values) == 1:
                            status = "PASS"
                            remarks = (
                                "Level 2 remained ambiguous; the detected "
                                "coding/combination uniquely resolves the RMS value."
                            )
                            recommended = c_rms_values[0]
                        else:
                            status = "REVIEW"
                            remarks = (
                                "Level 3 remains ambiguous; the coding/combination "
                                "does not uniquely identify one RMS value."
                            )
                            recommended = ""

                        analysis_rows.append({
                            "OMNI ID": omni_col,
                            "RMS ID": rms_col,
                            "Item Scope": "SAME IN BOTH",
                            "OMNI Value": "",
                            "RMS Value": " | ".join(c_rms_values),
                            "Alignment Statement": f"{omni_col} != {rms_col}",
                            "Validation Status": "FALSE",
                            "Unique RMS Value": " | ".join(c_rms_values),
                            "Derived Brand High": derived_display,
                            "#US LOC BRAND": brand_display,
                            "Coding / Combination": c_display,
                            "Rule Level": "LEVEL 3",
                            "Impacted Nankey Count": c_nankeys,
                            "Unique RMS Count": len(c_rms_values),
                            "Final Status": status,
                            "Remarks": remarks,
                        })

                        rule_rows.append({
                            "OMNI ID": omni_col,
                            "RMS ID": rms_col,
                            "Rule Level": "LEVEL 3",
                            "Derived Brand High": derived_display,
                            "#US LOC BRAND": brand_display,
                            "Coding / Combination": c_display,
                            "RMS Value": " | ".join(c_rms_values),
                            "Impacted Nankey Count": c_nankeys,
                            "Unique RMS Count": len(c_rms_values),
                            "Status": status,
                            "Recommended OMNI Value": recommended,
                            "Remarks": remarks,
                        })

    analysis_df = pd.DataFrame(analysis_rows, columns=ANALYSIS_COLUMNS)
    rule_df = pd.DataFrame(rule_rows, columns=RULE_COLUMNS)

    # If no differences were found, create an explicit trace row rather than
    # silently producing an empty analysis.
    if analysis_df.empty:
        analysis_df = pd.DataFrame([{
            col: (
                "No FALSE alignment records found for SAME IN BOTH scope."
                if col == "Remarks" else ""
            )
            for col in ANALYSIS_COLUMNS
        }])

    if rule_df.empty:
        rule_df = pd.DataFrame(columns=RULE_COLUMNS)

    return analysis_df, rule_df, {
        "item_scope_col": item_scope_col,
        "derived_col": derived_col,
        "brand_col": brand_col,
        "nankey_col": nankey_col,
        "coding_cols": coding_cols,
        "scoped_rows": len(scoped),
    }


# -----------------------------
# Excel output
# -----------------------------

def style_sheet(ws, freeze="A2"):
    ws.freeze_panes = freeze
    ws.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D1D5DB")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # Practical widths.
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells[:300]:
            value = clean_value(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 42)

    ws.auto_filter.ref = ws.dimensions


def write_dataframe_sheet(wb, title: str, df: pd.DataFrame):
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)

    # Header
    for c, col in enumerate(df.columns, start=1):
        ws.cell(1, c, col)

    # Data
    for r, row in enumerate(df.itertuples(index=False, name=None), start=2):
        for c, value in enumerate(row, start=1):
            if pd.isna(value):
                value = None
            ws.cell(r, c, value)

    style_sheet(ws)
    return ws


def remove_filters_and_unhide(ws):
    ws.auto_filter.ref = None
    # Remove any worksheet-level filter state and unhide rows.
    for row in ws.iter_rows():
        if row:
            row[0].parent.row_dimensions[row[0].row].hidden = False


def create_output(input_path: str, output_path: str):
    """
    Complete workbook build. Original OMNI IV and MAC sheets are preserved;
    Analysis and Rule are generated.
    """
    if not os.path.isfile(input_path):
        raise FileNotFoundError(f"Input workbook not found: {input_path}")

    # Validate workbook structure first.
    source_wb = load_workbook(input_path, read_only=False, data_only=False)
    if INPUT_OMNI_SHEET not in source_wb.sheetnames:
        raise ValueError("OMNI IV sheet is missing.")
    if INPUT_MAC_SHEET not in source_wb.sheetnames:
        raise ValueError("MAC sheet is missing.")
    source_wb.close()

    df, _, _ = read_omni_sheet(input_path)
    mappings = read_mac_sheet(input_path)

    # Build using actual workbook headers.
    analysis_df, rule_df, metadata = build_analysis_and_rules(df, mappings)

    # Load original workbook again to preserve OMNI IV and MAC.
    wb = load_workbook(input_path, read_only=False, data_only=False)

    # Delete any pre-existing generated sheets so output is exactly four sheets.
    for sheet in list(wb.sheetnames):
        if sheet not in [INPUT_OMNI_SHEET, INPUT_MAC_SHEET]:
            del wb[sheet]

    # Remove filters from preserved input sheets as requested.
    remove_filters_and_unhide(wb[INPUT_OMNI_SHEET])
    remove_filters_and_unhide(wb[INPUT_MAC_SHEET])

    write_dataframe_sheet(wb, "Analysis", analysis_df)
    write_dataframe_sheet(wb, "Rule", rule_df)

    # Enforce exact sheet order.
    wb._sheets = [
        wb[INPUT_OMNI_SHEET],
        wb[INPUT_MAC_SHEET],
        wb["Analysis"],
        wb["Rule"],
    ]

    wb.save(output_path)
    wb.close()

    return metadata, analysis_df, rule_df


# -----------------------------
# Dark GUI
# -----------------------------

class RMSOmniApp:
    BG = "#111827"
    PANEL = "#1F2937"
    PANEL_2 = "#0F172A"
    TEXT = "#F9FAFB"
    MUTED = "#9CA3AF"
    ACCENT = "#2563EB"
    ACCENT_HOVER = "#1D4ED8"
    SUCCESS = "#10B981"
    ERROR = "#EF4444"

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("RMS vs OMNI Alignment Validator")
        self.root.geometry("760x460")
        self.root.minsize(700, 420)
        self.root.configure(bg=self.BG)

        self.input_path = tk.StringVar(value="")
        self.output_path = tk.StringVar(value="")
        self.status = tk.StringVar(value="Ready")
        self.progress_value = tk.DoubleVar(value=0)

        self._build_styles()
        self._build_ui()

    def _build_styles(self):
        style = ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except tk.TclError:
            pass

        style.configure(
            "Dark.Horizontal.TProgressbar",
            troughcolor=self.PANEL_2,
            background=self.ACCENT,
            bordercolor=self.PANEL_2,
            lightcolor=self.ACCENT,
            darkcolor=self.ACCENT,
        )

        style.configure(
            "Dark.TButton",
            background=self.PANEL,
            foreground=self.TEXT,
            borderwidth=0,
            padding=(18, 10),
            font=("Segoe UI", 10, "bold"),
        )
        style.map(
            "Dark.TButton",
            background=[("active", self.ACCENT)],
            foreground=[("active", "#FFFFFF")],
        )

    def _build_ui(self):
        title = tk.Label(
            self.root,
            text="RMS vs OMNI Alignment Validator",
            bg=self.BG,
            fg=self.TEXT,
            font=("Segoe UI Semibold", 20),
        )
        title.pack(pady=(28, 4))

        subtitle = tk.Label(
            self.root,
            text="Dynamic Excel validation for reusable client workbooks",
            bg=self.BG,
            fg=self.MUTED,
            font=("Segoe UI", 10),
        )
        subtitle.pack(pady=(0, 22))

        card = tk.Frame(self.root, bg=self.PANEL, highlightthickness=1,
                        highlightbackground="#374151")
        card.pack(fill="both", expand=True, padx=34, pady=0)

        # File upload
        upload_frame = tk.Frame(card, bg=self.PANEL)
        upload_frame.pack(fill="x", padx=24, pady=(24, 14))

        tk.Label(
            upload_frame, text="FILE UPLOAD", bg=self.PANEL, fg=self.MUTED,
            font=("Segoe UI Semibold", 9)
        ).pack(anchor="w")

        file_row = tk.Frame(upload_frame, bg=self.PANEL)
        file_row.pack(fill="x", pady=(8, 0))

        self.file_entry = tk.Entry(
            file_row, textvariable=self.input_path, state="readonly",
            readonlybackground=self.PANEL_2, fg=self.TEXT,
            insertbackground=self.TEXT, relief="flat",
            font=("Segoe UI", 10)
        )
        self.file_entry.pack(side="left", fill="x", expand=True, ipady=9)

        self.browse_btn = ttk.Button(
            file_row, text="Browse", style="Dark.TButton",
            command=self.browse_file
        )
        self.browse_btn.pack(side="left", padx=(10, 0))

        # Action buttons
        actions = tk.Frame(card, bg=self.PANEL)
        actions.pack(fill="x", padx=24, pady=14)

        self.clear_btn = ttk.Button(
            actions, text="Clear", style="Dark.TButton",
            command=self.clear
        )
        self.clear_btn.pack(side="left")

        self.run_btn = tk.Button(
            actions, text="Run Validation", command=self.run_validation,
            bg=self.ACCENT, fg="#FFFFFF", activebackground=self.ACCENT_HOVER,
            activeforeground="#FFFFFF", relief="flat", bd=0,
            padx=22, pady=11, font=("Segoe UI Semibold", 10),
            cursor="hand2",
        )
        self.run_btn.pack(side="right")

        # Status
        status_frame = tk.Frame(card, bg=self.PANEL)
        status_frame.pack(fill="x", padx=24, pady=(10, 10))

        tk.Label(
            status_frame, text="STATUS", bg=self.PANEL, fg=self.MUTED,
            font=("Segoe UI Semibold", 9)
        ).pack(anchor="w")

        self.status_label = tk.Label(
            status_frame, textvariable=self.status, bg=self.PANEL,
            fg=self.TEXT, font=("Segoe UI", 10)
        )
        self.status_label.pack(anchor="w", pady=(6, 8))

        self.progress = ttk.Progressbar(
            status_frame, variable=self.progress_value, maximum=100,
            mode="determinate", style="Dark.Horizontal.TProgressbar"
        )
        self.progress.pack(fill="x")

        # Output
        output_frame = tk.Frame(card, bg=self.PANEL)
        output_frame.pack(fill="x", padx=24, pady=(20, 24))

        tk.Label(
            output_frame, text="OUTPUT", bg=self.PANEL, fg=self.MUTED,
            font=("Segoe UI Semibold", 9)
        ).pack(anchor="w")

        out_row = tk.Frame(output_frame, bg=self.PANEL)
        out_row.pack(fill="x", pady=(8, 0))

        self.output_entry = tk.Entry(
            out_row, textvariable=self.output_path, state="readonly",
            readonlybackground=self.PANEL_2, fg=self.TEXT,
            relief="flat", font=("Segoe UI", 9)
        )
        self.output_entry.pack(side="left", fill="x", expand=True, ipady=8)

        self.open_btn = tk.Button(
            out_row, text="Open Output", command=self.open_output,
            bg=self.PANEL_2, fg=self.TEXT, activebackground="#374151",
            activeforeground="#FFFFFF", relief="flat", bd=0,
            padx=16, pady=9, font=("Segoe UI Semibold", 9),
            state="disabled", cursor="hand2",
        )
        self.open_btn.pack(side="left", padx=(10, 0))

    def browse_file(self):
        path = filedialog.askopenfilename(
            title="Select RMS vs OMNI Excel Workbook",
            filetypes=[
                ("Excel Workbook", "*.xlsx"),
                ("Excel Macro-Enabled Workbook", "*.xlsm"),
                ("All Files", "*.*"),
            ],
        )
        if path:
            self.input_path.set(path)
            self.output_path.set("")
            self.status.set("Workbook selected. Ready to validate.")
            self.progress_value.set(0)
            self.open_btn.config(state="disabled")

    def clear(self):
        self.input_path.set("")
        self.output_path.set("")
        self.status.set("Ready")
        self.progress_value.set(0)
        self.open_btn.config(state="disabled")

    def set_busy(self, busy: bool):
        state = "disabled" if busy else "normal"
        self.browse_btn.config(state=state)
        self.clear_btn.config(state=state)
        self.run_btn.config(state=state)

    def run_validation(self):
        input_path = self.input_path.get().strip()
        if not input_path:
            messagebox.showwarning(
                "Input Required", "Please select an Excel workbook first."
            )
            return

        if not os.path.isfile(input_path):
            messagebox.showerror("File Error", "The selected workbook no longer exists.")
            return

        desktop = Path.home() / "Desktop"
        desktop.mkdir(parents=True, exist_ok=True)
        output_path = desktop / OUTPUT_FILENAME

        self.set_busy(True)
        self.open_btn.config(state="disabled")
        self.progress_value.set(10)
        self.status.set("Validating workbook...")

        worker = threading.Thread(
            target=self._worker,
            args=(input_path, str(output_path)),
            daemon=True,
        )
        worker.start()

    def _worker(self, input_path: str, output_path: str):
        try:
            self.root.after(0, lambda: self.progress_value.set(25))
            self.root.after(0, lambda: self.status.set("Reading workbook structure..."))

            metadata, analysis_df, rule_df = create_output(
                input_path, output_path
            )

            self.root.after(0, lambda: self.progress_value.set(100))
            self.root.after(
                0,
                lambda: self._success(
                    output_path, len(analysis_df), len(rule_df)
                ),
            )
        except Exception as exc:
            detail = f"{exc}\n\n{traceback.format_exc(limit=3)}"
            self.root.after(0, lambda: self._failure(detail))

    def _success(self, output_path: str, analysis_count: int, rule_count: int):
        self.set_busy(False)
        self.output_path.set(output_path)
        self.status.set(
            f"Completed — {rule_count} rule record(s), "
            f"{analysis_count} analysis record(s)."
        )
        self.status_label.config(fg=self.SUCCESS)
        self.open_btn.config(state="normal")
        messagebox.showinfo(
            "Validation Complete",
            "The RMS vs OMNI alignment workbook was created successfully.\n\n"
            f"Output:\n{output_path}"
        )

    def _failure(self, detail: str):
        self.set_busy(False)
        self.progress_value.set(0)
        self.status.set("Validation failed.")
        self.status_label.config(fg=self.ERROR)
        messagebox.showerror("Validation Error", detail)

    def open_output(self):
        path = self.output_path.get().strip()
        if not path or not os.path.isfile(path):
            messagebox.showwarning(
                "Output Not Found", "The generated output file is not available."
            )
            return

        try:
            os.startfile(path)  # Windows
        except AttributeError:
            messagebox.showerror(
                "Windows Required",
                "Opening the output directly is supported on Windows."
            )


def main():
    root = tk.Tk()
    app = RMSOmniApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
